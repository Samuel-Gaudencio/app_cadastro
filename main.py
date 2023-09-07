import tkinter
from tkinter import END
import customtkinter as ctk
import openpyxl

e_cod = None
e_name = None
e_valor = None
e_quant = None
e_repre = None
e_class = None


def limpar_campos():
    e_cod.delete(0, END)
    e_name.delete(0, END)
    e_valor.delete(0, END)
    e_quant.delete(0, END)
    e_repre.delete(0, END)
    e_class.delete(0, END)


def cadastrar_produto():
    cod = e_cod.get()
    name = e_name.get()
    valor = e_valor.get()
    quant = e_quant.get()
    repre = e_repre.get()
    classi = e_class.get()

    fichario = openpyxl.load_workbook('registro.xlsx')
    folha = fichario.active

    folha.cell(column=1, row=folha.max_row+1, value=cod)
    folha.cell(column=2, row=folha.max_row, value=name)
    folha.cell(column=3, row=folha.max_row, value=valor)
    folha.cell(column=4, row=folha.max_row, value=quant)
    folha.cell(column=5, row=folha.max_row, value=repre)
    folha.cell(column=6, row=folha.max_row, value=classi)

    if name == "" or cod == "" or valor == "" or quant == "" or repre == "" or classi == "":
        tkinter.messagebox.showwarning('Sistema', "ERROR\nPor favor preencha todos os campos!!")
    else:
        fichario.save(r'registro.xlsx')
        tkinter.messagebox.showinfo('Sistema', "Produto cadastrado com sucesso!!")


def janela_cadastro():
    global e_cod, e_name, e_valor, e_quant, e_repre, e_class

    cadastro = ctk.CTkToplevel(root)
    cadastro.geometry('600x400')
    cadastro.title('Cadastro de Produtos')
    cadastro.maxsize(600, 400)
    cadastro.minsize(600, 400)
    cadastro.focus_force()
    cadastro.grab_set()
    ctk.set_appearance_mode('dark')
    ctk.set_default_color_theme('blue')

    titulo = ctk.CTkLabel(cadastro, text='Cadastro de Produtos', text_color='#fff', font=('Montserrat', 25, 'bold'))
    titulo.pack(pady=20)

    label_cod = ctk.CTkLabel(cadastro, text='Código do Produto:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_cod.place(x=10, y=70)

    e_cod = ctk.CTkEntry(cadastro, placeholder_text='Código do Produto', text_color='#fff', font=('Montserrat', 12, 'bold'),
                         placeholder_text_color='#fff', width=250, fg_color='transparent', border_width=1)
    e_cod.place(x=10, y=95)

    label_name = ctk.CTkLabel(cadastro, text='Nome do Produto:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_name.place(x=300, y=70)

    e_name = ctk.CTkEntry(cadastro, placeholder_text='Nome do Produto', text_color='#fff',
                          font=('Montserrat', 12, 'bold'),
                          placeholder_text_color='#fff', width=250, fg_color='transparent', border_width=1)

    e_name.place(x=300, y=95)

    label_valor = ctk.CTkLabel(cadastro, text='Valor do Produto:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_valor.place(x=10, y=150)

    e_valor = ctk.CTkEntry(cadastro, placeholder_text='R$: ', text_color='#fff', font=('Montserrat', 12, 'bold'),
                           placeholder_text_color='#fff', width=250, fg_color='transparent', border_width=1)

    e_valor.place(x=10, y=175)

    label_quant = ctk.CTkLabel(cadastro, text='Quantidade do Produto:', text_color='#fff',
                               font=('Montserrat', 16, 'bold'))
    label_quant.place(x=300, y=150)

    e_quant = ctk.CTkEntry(cadastro, placeholder_text='Quantidade do Produto', text_color='#fff',
                           font=('Montserrat', 12, 'bold'),
                           placeholder_text_color='#fff', width=250, fg_color='transparent', border_width=1)

    e_quant.place(x=300, y=175)

    label_repre = ctk.CTkLabel(cadastro, text='Representante do Produto:', text_color='#fff',
                               font=('Montserrat', 16, 'bold'))
    label_repre.place(x=10, y=230)

    e_repre = ctk.CTkEntry(cadastro, placeholder_text='Representante do Produto', text_color='#fff',
                           font=('Montserrat', 12, 'bold'),
                           placeholder_text_color='#fff', width=250, fg_color='transparent', border_width=1)

    e_repre.place(x=10, y=255)

    label_class = ctk.CTkLabel(cadastro, text='Tipo do Produto:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_class.place(x=300, y=230)

    e_class = ctk.CTkEntry(cadastro, placeholder_text='Tipo do Produto', text_color='#fff',
                           font=('Montserrat', 12, 'bold'),
                           placeholder_text_color='#fff', width=250, fg_color='transparent', border_width=1)

    e_class.place(x=300, y=255)

    b_cadas = ctk.CTkButton(cadastro, text='CADASTRAR', text_color='#fff', fg_color='#C69749',
                            font=('Montserrat', 20, 'bold'),
                            hover_color='#FF6400', command=cadastrar_produto)
    b_cadas.place(x=130, y=330)

    b_limpar = ctk.CTkButton(cadastro, text='LIMPAR', text_color='#fff', fg_color='#C69749',
                             font=('Montserrat', 20, 'bold'),
                             hover_color='#FF6400', command=limpar_campos)
    b_limpar.place(x=320, y=330)


def janela_estoque():
    estoque = ctk.CTkToplevel(root)
    estoque.geometry('600x350')
    estoque.title('Estoque')
    estoque.maxsize(600, 350)
    estoque.minsize(600, 350)
    estoque.focus_force()
    estoque.grab_set()
    ctk.set_appearance_mode('dark')
    ctk.set_default_color_theme('blue')

    fichario = openpyxl.load_workbook('registro.xlsx')
    folha = fichario.active
    lista = []

    for row in folha.iter_cols(values_only=True):
        lista.append(list(row))

    codigo = lista[0]
    produtos = lista[1]
    value = lista[2]
    quant_pro = lista[3]
    repre_pro = lista[4]
    tipo_pro = lista[5]

    def procurar():
        nome_prod = m_name.get()
        posicao = produtos.index(nome_prod)
        cod.configure(text=codigo[posicao])
        valor.configure(text=value[posicao])
        quant.configure(text=quant_pro[posicao])
        repre.configure(text=repre_pro[posicao])
        tipo.configure(text=tipo_pro[posicao])

    titulo = ctk.CTkLabel(estoque, text='Estoque', text_color='#fff', font=('Montserrat', 28, 'bold'))
    titulo.pack(pady=20)

    label_name = ctk.CTkLabel(estoque, text='Nome do Produto:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_name.place(x=30, y=90)

    m_name = ctk.CTkOptionMenu(estoque, values=produtos, text_color='#fff', fg_color='#C69749', button_color='#C69749',
                               button_hover_color='#C69749', font=('Montserrat', 16), width=200)
    m_name.place(x=180, y=90)

    buttun = ctk.CTkButton(estoque, text='Procurar', text_color='#fff', fg_color='#C69749', hover_color='#FF6400',
                           font=('Montserrat', 16, 'bold'), command=procurar)
    buttun.place(x=430, y=90)

    label_cod = ctk.CTkLabel(estoque, text='Codigo do Produto:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_cod.place(x=30, y=130)

    cod = ctk.CTkLabel(estoque, text=codigo[0], text_color='#fff', font=('Montserrat', 16, 'bold'))
    cod.place(x=190, y=130)

    label_valor = ctk.CTkLabel(estoque, text='Valor do Produto R$:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_valor.place(x=30, y=170)

    valor = ctk.CTkLabel(estoque, text=value[0], text_color='#fff', font=('Montserrat', 16, 'bold'))
    valor.place(x=200, y=170)

    label_quant = ctk.CTkLabel(estoque, text='Quantidade do Produto:', text_color='#fff',
                               font=('Montserrat', 16, 'bold'))
    label_quant.place(x=30, y=210)

    quant = ctk.CTkLabel(estoque, text=quant_pro[0], text_color='#fff', font=('Montserrat', 16, 'bold'))
    quant.place(x=220, y=210)

    label_repre = ctk.CTkLabel(estoque, text='Representante do Produto:', text_color='#fff',
                               font=('Montserrat', 16, 'bold'))
    label_repre.place(x=30, y=250)

    repre = ctk.CTkLabel(estoque, text=repre_pro[0], text_color='#fff', font=('Montserrat', 16, 'bold'))
    repre.place(x=245, y=250)

    label_tipo = ctk.CTkLabel(estoque, text='Tipo do Produto:', text_color='#fff', font=('Montserrat', 16, 'bold'))
    label_tipo.place(x=30, y=290)

    tipo = ctk.CTkLabel(estoque, text=tipo_pro[0], text_color='#fff', font=('Montserrat', 16, 'bold'))
    tipo.place(x=170, y=290)


root = ctk.CTk()
root.geometry('400x200')
root.maxsize(400, 200)
root.minsize(400, 200)

title = ctk.CTkLabel(root, text='Bem-Vindo ao Painel da Loja', text_color='#fff', font=('Montserrat', 25, 'bold'))
title.pack(pady=30)

botao_cadas = ctk.CTkButton(root, text='Cadastrar Produtos', text_color='#fff', font=('Montserrat', 20, 'bold'),
                            fg_color='#C69749',  hover_color='#FF6400', command=janela_cadastro, width=200)
botao_cadas.place(x=20, y=100)

botao_check = ctk.CTkButton(root, text='Estoque', text_color='#fff', font=('Montserrat', 20, 'bold'), fg_color='#C69749',
                            hover_color='#FF6400', command=janela_estoque, width=150)
botao_check.place(x=240, y=100)

root.mainloop()
